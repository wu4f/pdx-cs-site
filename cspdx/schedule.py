"""Generate the CS course schedule spreadsheet from Banner SSB.

Fetches the 3 most recent terms, writes each to a separate sheet,
and saves the result to the given path.
"""
from __future__ import annotations
import html
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

BASE_URL = "https://app.banner.pdx.edu/StudentRegistrationSsb/ssb"

_HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; PSU-schedule-fetcher/1.0)",
    "Accept": "application/json, text/javascript, */*; q=0.01",
    "X-Requested-With": "XMLHttpRequest",
}

_COLUMNS = [
    ("CRN",        "courseReferenceNumber", 8),
    ("Course",     None,                    10),
    ("Section",    "sequenceNumber",        8),
    ("Title",      "courseTitle",           36),
    ("Credits",    "creditHours",           8),
    ("Days",       None,                    7),
    ("Time",       None,                    13),
    ("Instructor", None,                    24),
]

_HDR_FILL  = PatternFill("solid", fgColor="003366")
_HDR_FONT  = Font(color="FFFFFF", bold=True)
_EVEN_FILL = PatternFill("solid", fgColor="EEF2FF")


def _get_terms(n: int = 3) -> list[dict]:
    resp = requests.get(
        f"{BASE_URL}/classSearch/getTerms",
        params={"searchTerm": "", "offset": 1, "max": 100},
        headers=_HEADERS,
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()[:n]


def _establish_session(term_code: str) -> requests.Session:
    session = requests.Session()
    session.headers.update(_HEADERS)
    session.get(f"{BASE_URL}/classSearch/classSearch", timeout=30)
    session.get(
        f"{BASE_URL}/term/search",
        params={
            "mode": "search",
            "term": term_code,
            "studyPath": "",
            "studyPathText": "",
            "startDatepicker": "",
            "endDatepicker": "",
        },
        timeout=30,
    )
    return session


def _fetch_courses(session: requests.Session, term_code: str, subject: str = "CS", page_size: int = 500) -> list[dict]:
    courses: list[dict] = []
    offset = 0
    while True:
        resp = session.get(
            f"{BASE_URL}/searchResults/searchResults",
            params={
                "txt_subject": subject,
                "txt_term": term_code,
                "startDatepicker": "",
                "endDatepicker": "",
                "pageOffset": offset,
                "pageMaxSize": page_size,
                "sortColumn": "subjectDescription",
                "sortDirection": "asc",
            },
            timeout=60,
        )
        resp.raise_for_status()
        data = resp.json()
        if not data.get("success") or not data.get("data"):
            break
        courses.extend(data["data"])
        total = data.get("totalCount", 0)
        offset += page_size
        if offset >= total:
            break
    return courses


def _meeting_info(course: dict) -> tuple[str, str]:
    meetings = course.get("meetingsFaculty", [])
    if not meetings:
        return "", ""
    mt = meetings[0].get("meetingTime", {}) or {}
    day_map = [
        ("monday", "M"), ("tuesday", "T"), ("wednesday", "W"),
        ("thursday", "R"), ("friday", "F"), ("saturday", "S"), ("sunday", "U"),
    ]
    days = "".join(abbr for key, abbr in day_map if mt.get(key))
    begin, end = mt.get("beginTime") or "", mt.get("endTime") or ""
    time_str = ""
    if begin and end and len(begin) == 4 and len(end) == 4:
        time_str = f"{begin[:2]}:{begin[2:]}-{end[:2]}:{end[2:]}"
    return days, time_str


def _instructor(course: dict) -> str:
    faculty = course.get("faculty") or []
    primary = next((f for f in faculty if f.get("primaryIndicator")), faculty[0] if faculty else None)
    return (primary.get("displayName") or "").strip() if primary else ""


def _row(course: dict) -> list:
    days, time_str = _meeting_info(course)
    return [
        course.get("courseReferenceNumber", ""),
        f"{course.get('subject', '')} {course.get('courseNumber', '')}".strip(),
        course.get("sequenceNumber", ""),
        html.unescape(course.get("courseTitle", "") or ""),
        course.get("creditHours", ""),
        days,
        time_str,
        _instructor(course),
    ]


def _write_sheet(wb: Workbook, courses: list[dict], sheet_name: str) -> None:
    ws = wb.create_sheet(title=sheet_name[:31])
    ws.append([col[0] for col in _COLUMNS])
    for cell in ws[1]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18
    for row_idx, course in enumerate(courses, start=2):
        ws.append(_row(course))
        if row_idx % 2 == 0:
            for cell in ws[row_idx]:
                cell.fill = _EVEN_FILL
    for col_idx, (_, _, width) in enumerate(_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def generate_schedule(out_path: Path) -> None:
    """Fetch Banner schedule for the 3 most recent terms and save to out_path."""
    print("[schedule] fetching available terms...", flush=True)
    terms = _get_terms(3)
    if not terms:
        raise RuntimeError("could not retrieve terms from Banner")

    descriptions = [t["description"].replace(" (View Only)", "").strip() for t in terms]
    print(f"[schedule] terms: {', '.join(descriptions)}", flush=True)

    wb = Workbook()
    wb.remove(wb.active)

    for term in terms:
        code = term["code"]
        desc = term["description"].replace(" (View Only)", "").strip()
        print(f"[schedule] [{desc}] establishing session...", flush=True)
        session = _establish_session(code)
        print(f"[schedule] [{desc}] fetching CS courses...", flush=True)
        courses = _fetch_courses(session, code)
        print(f"[schedule] [{desc}] {len(courses)} sections found", flush=True)
        _write_sheet(wb, courses, desc)

    if not wb.sheetnames:
        raise RuntimeError("no data retrieved for any term")

    wb.active = wb.worksheets[0]
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[schedule] saved -> {out_path}", flush=True)
